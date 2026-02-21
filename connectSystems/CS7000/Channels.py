import os
import csv
import sys
import openpyxl
from connectSystems.CS7000.DigitalContacts import DigitalContacts
from openpyxl import Workbook
from hashlib import sha256
from decimal import Decimal
class Channels:

    def __init__(self, input_file, output_file, includeAnalogChannels):
        self.input_file = input_file
        self.output_file = output_file
        self._UhfChannels = {}
        self._VhfChannels = {}

        # Track where firmware analog and digital channels end
        # Note: In order for this to work you must load firmware channels file
        self._endFirmwareAnalog = 0
        self._endFirmwareDigital = 0

        # Setup header and default row records
        self._SetArrays()
        self._channelRowsAnalog = []
        self._channelRowsDigital = []
        self._fileType = ''
        self._fileVersion = ''
        self.load(self.input_file)

        self._includeAnalogChannels = includeAnalogChannels

        # Track rows we have written to worksheet
        self._analogRowsWritten = 0
        self._digitalRowsWritten = 0

    def channel_type(self, e : list):
        bw = e[3]
        if ( bw == "12.5"):
            return "Analog"
        elif ( bw == "25"):
            return "Analog"
        else:
            return "DMR"

    # If channel is from firmware return True otherwise false

    def isFirmwareChannel(self, e : list):
        if self.channel_type(e) == "Analog":
            if int(e[0]) <= self._endFirmwareAnalog:
                return True
            else:
                return False
        elif self.channel_type(e) == "DMR":
            if int(e[0]) <= self._endFirmwareDigital:
                return True
            else:
                return False

    def get_all(self):
        analog = self.get_all_analog()
        digital = self.get_all_digital()
        combined = [*analog, *digital]
        return combined


    def get_all_analog(self):
        return list(self._channelRowsAnalog)

    def get_all_digital(self):
        return list(self._channelRowsDigital)

    def find_analog_by_name(self, name):
        return [c for c in self._channelRowsAnalog if c[1] == name]

    def find_digital_by_name(self, name):
        return [c for c in self._channelRowsDigital if c[1] == name]

    def remove(self, ch : list):
        if self.channel_type(ch) == "Analog":
            self.remove_analog(ch[1])
        elif self.channel_type(ch) == "DMR":
            self.remove_digital(ch[1])

    def remove_analog(self, name):
        self._channelRowsAnalogTemp = [c for c in self._channelRowsAnalog if c[1] != name]
        self._channelRowsAnalog = self._channelRowsAnalogTemp
        if name in self_UhfChannels:
            del self._UhfChannels[name]
        elif name in self._VhfChannels:
            del self._VhfChannels[name]

    def remove_digital(self, name):
        self._channelRowsDigitalTemp = [c for c in self._channelRowsDigital if c[1] != name]
        self._channelRowsDigital = self._channelRowsDigitalTemp
        if name in self._UhfChannels:
            del self._UhfChannels[name]
        elif name in self._VhfChannels:
            del self._vhfChannels[name]

    def update_contact_name(self, old_alias, new_alias):
        for ch in self._channelRowsDigital:
            if ch[19] == old_alias:
                ch[19] = new_alias

    def getVhfChannels(self):
        return (self._VhfChannels.copy())

    def getUhfChannels(self):
        return (self._UhfChannels.copy())

    def getNumberChannels(self):
        return len(self._channelRowsAnalog)+len(self._channelRowsDigital)

    def load(self,input_file):
        self.input_file = input_file
        self._DetermineFileType()

        # Open CSV input
        with open(input_file, mode='r', newline='') as infile:
            reader = csv.reader(infile)
            # Get row number of number of elements in the channelRowsAnalog and channelRowsDigital
            rowNum_dmr = len(self._channelRowsDigital)
            rowNum_analog = len(self._channelRowsAnalog)

            # ---------------------------------------------------------
            # PROCESS CSV ROWS
            # ---------------------------------------------------------
            readHeaderRow = 0
            for row in reader:
                outputRowDMR = self._defaultRow_dmr[:]
                outputRowAnalog = self._defaultRow_analog[:]
                if readHeaderRow == 0:
                    readHeaderRow = 1
                    continue

                # Check if file type is CS7000 analog or digital channels.  If so write the row from the input CSV
                if (self._fileType == 'CS7000_analog_channels'):
                    self._channelRowsAnalog.append(row)
                else:
                    if (self._fileType == 'CS7000_digital_channels'):
                        self._channelRowsDigital.append(row)
                    else:
                        # Anytone
                        channelName = row[1]
                        mode = row[4]

                        # Normalize mode
                        if mode == "A-Analog":
                            mode = "FM"
                        if mode == "D-Digital":
                            mode = "DMR"

                        power = row[5]
                        if power == "Turbo":
                            power = "High"
                        if power == "Mid":
                            power = "Low"

                        channelSpacing = row[6]
                        channelSpacing = 25.0 if channelSpacing == "25K" else 12.5

                        tx_freq = row[3]
                        rx_freq = row[2]

                        # Check if VHF and if so add to list of VHF channels
                        if (float(tx_freq) <= 300.0) and (float(rx_freq) <= 300):
                            self._VhfChannels[channelName] = True

                        ptt_prohibit = row[24]
                        dmr_tx_permit = row[13]

                        if dmr_tx_permit == "Off":
                            dmr_tx_permit = "Always"
                        if dmr_tx_permit == "ChannelFree":
                            dmr_tx_permit = "Channel Idle"

                        if mode == "DMR":
                            call_alias = row[9]
                            timeslot = row[21]
                            colorCode = row[20]
                        else:
                            ctcss_decode = row[7]
                            ctcss_encode = row[8]

                        # ---------------------------------------------------------
                        # BUILD OUTPUT ROWS
                        # ---------------------------------------------------------
                        if mode == "DMR":
                            outputRowDMR[0] = rowNum_dmr
                            outputRowDMR[1] = channelName
                            outputRowDMR[2] = rowNum_dmr
                            outputRowDMR[3] = colorCode
                            outputRowDMR[4] = "Slot 1" if Decimal(timeslot) == 1 else "Slot 2"
                            outputRowDMR[7] = ptt_prohibit
                            outputRowDMR[11] = rx_freq
                            outputRowDMR[17] = tx_freq
                            outputRowDMR[22] = dmr_tx_permit
                            outputRowDMR[19] = "None" if call_alias == "-NULL-" else call_alias
                            outputRowDMR[21] = power

                        else:
                            outputRowAnalog[0] = rowNum_analog
                            outputRowAnalog[1] = channelName
                            outputRowAnalog[3] = channelSpacing
                            outputRowAnalog[7] = ptt_prohibit
                            outputRowAnalog[11] = rx_freq
                            outputRowAnalog[18] = tx_freq

                            if ctcss_encode != "Off":
                                outputRowAnalog[19] = "CTCSS"
                                outputRowAnalog[20] = ctcss_encode
                            else:
                                outputRowAnalog[19] = "NONE"
                                outputRowAnalog[20] = "NONE"

                            if ctcss_decode != "Off":
                                outputRowAnalog[12] = "CTCSS"
                                outputRowAnalog[13] = ctcss_decode
                            else:
                                outputRowAnalog[12] = "NONE"
                                outputRowAnalog[13] = "NONE"

                            outputRowAnalog[22] = power

                        # Input row has been converted
                        # Determine channel type and append to list
                        if ( mode == "DMR"):
                            self._channelRowsDigital.append(outputRowDMR[:])
                            rowNum_dmr += 1
                        else:
                            if ( mode == "FM"):
                                self._channelRowsAnalog.append(outputRowAnalog[:])
                                rowNum_analog += 1
        if (self._fileType == "CS7000_analog_channels"):
            self._endFirmwareAnalog = len(self._channelRowsAnalog)
        elif (self._fileType == "CS7000_digital_channels"):
            self._endFirmwareDigital = len(self._channelRowsDigital)

    def whereEndFirmware(self):
        a = self._endFirmwareAnalog
        print(f"Firmware end of Analog Channels = {a}  Digital Channels = {self._endFirmwareDigital}")

    def Convert(self):
        if self._fileType in ("Anytone", "CS7000_analog_channels", "CS7000_digital_channels"):
            self.writeToSpreadsheet(True)
            self.LoadChannelNames(self.output_file)
            return (self._UhfChannels.copy())
        if self._fileType == "ERROR":
            return (-1)

    def ConvertDirectMode(self, contacts : DigitalContacts):
        # Update Digital Channels to use Direct mode output vice Table lookup 
        numAliasNotFound = 0
        aliasNotFound = ""
        if self._fileType in ("Anytone", "CS7000_analog_channels", "CS7000_digital_channels"):

            for row in self._channelRowsDigital:
                # 1. Get alias
                alias = row[19]
                if alias == "0":
                    continue

                # 2. Lookup contact info
                try:
                    result = contacts.getContact(alias)
                except:
                    aliasNotFound += f"{alias}\n"
                    numAliasNotFound += 1
                    pass

                # Defensive check in case contact lookup fails
                if not result or len(result) < 3:
                    raise ValueError(f"Invalid contact lookup for alias '{alias}': {result}")

                # 3. Extract required fields
                digitalContact = result[1]
                contactType    = result[3]   # second element

                if (contactType == "Private Call"):
                    contactType = "Private"
                else:
                    contactType = "Group"

                # 4. Update row fields
                row[19] = digitalContact     # overwrite alias with digitalContact
                row[-2] = "Direct"           # second-to-last element
                row[-1] = contactType        # last element
            # Now that we have updated every row call method to output codeplug
            if (numAliasNotFound > 0):
                #update_debug_output("Did not find {numAliasNotFond} talkgroups in TalkGroups.CSV\n{aliasNotFound}")
                print(f"Did not find {numAliasNotFound} talkgroups in TalkGroups.CSV\n{aliasNotFound} {type(aliasNotFound)}")

            return (self.Convert())
        if self._fileType == "ERROR":
            return (-1)

    def _find_first_empty_row(self, ws, col=1):
        row = 2
        while True:
            value = ws.cell(row=row, column=col).value
            if value is None or str(value).strip() == "":
                return row
            row += 1

    def writeToSpreadsheet(self, appendToSpreadsheet=False):
        output_file = self.output_file
        writeHeaderRows = False
        # ---------------------------------------------------------
        # OPEN OR CREATE WORKBOOK
        # ---------------------------------------------------------
        if appendToSpreadsheet and os.path.exists(output_file):
            # File exists → append to it:
            writeHeaderRows = False
            workbook = openpyxl.load_workbook(output_file)
        else:
            # File does NOT exist OR appendToSpreadsheet=False → create new workbook
            writeHeaderRows = True
            workbook = Workbook()
            # Remove default sheet
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            # Create sheets
            workbook.create_sheet("Analog Channels")
            workbook.create_sheet("Digital Channels")
            workbook.create_sheet("Easy Trunking Channels")
            workbook.create_sheet("Easy Trunking Pool")

        analogWorksheet = workbook["Analog Channels"]
        digitalWorksheet = workbook["Digital Channels"]
        easyChannelsWorksheet = workbook["Easy Trunking Channels"]
        easyPoolWorksheet = workbook["Easy Trunking Pool"]

        # ---------------------------------------------------------
        # DETERMINE FIRST EMPTY ROW TO START WRITING CHANNELS
        # ---------------------------------------------------------
        if (self._analogRowsWritten == 0 or self._digitalRowsWritten == 0):
            rowNum_dmr = 2   # openpyxl is 1‑based; row 1 = header
            rowNum_analog = 2
        else:
            rowNum_analog = self._find_first_empty_row(analogWorksheet)
            rowNum_dmr = self._find_first_empty_row(digitalWorksheet)

        # ---------------------------------------------------------
        # WRITE HEADERS (only if not appending)
        # ---------------------------------------------------------
        if not appendToSpreadsheet or writeHeaderRows:
            for col, val in enumerate(self._header_dmr, start=1):
                digitalWorksheet.cell(row=1, column=col, value=val)

            for col, val in enumerate(self._header_analog, start=1):
                analogWorksheet.cell(row=1, column=col, value=val)

            for col, val in enumerate(self._header_easyTrunkingChannels, start=1):
                easyChannelsWorksheet.cell(row=1, column=col, value=val)

            for col, val in enumerate(self._header_easyTrunkingPool, start=1):
                easyPoolWorksheet.cell(row=1, column=col, value=val)

        # ---------------------------------------------------------
        # WRITE TO SHEETS
        # ---------------------------------------------------------

        # Loop through DMR channels
        for outputRowDMR in self._channelRowsDigital:
            rx_freq = Decimal(str(outputRowDMR[11]).strip())
            if rx_freq >= 400:
                for col, val in enumerate(outputRowDMR, start=1):
                    digitalWorksheet.cell(row=rowNum_dmr, column=col, value=val)
                # We must update column A "No." and C "Digital ID [Per Each Channel]"
                # Note the value is one less than the row number
                digitalWorksheet.cell(row=rowNum_dmr, column=1, value=rowNum_dmr-1)
                digitalWorksheet.cell(row=rowNum_dmr, column=3, value=rowNum_dmr-1)
                rowNum_dmr += 1
        for outputRowAnalog in self._channelRowsAnalog:
            rx_freq = Decimal(str(outputRowAnalog[11]).strip())
            if rx_freq >= 400 and self._includeAnalogChannels:
                for col, val in enumerate(outputRowAnalog, start=1):
                    analogWorksheet.cell(row=rowNum_analog, column=col, value=val)
                # We must update column A "No."
                analogWorksheet.cell(row=rowNum_analog, column=1, value=rowNum_analog-1)
                rowNum_analog += 1

        # Save workbook
        workbook.save(output_file)

    def _DetermineFileType(self):

        with open(self.input_file, mode='r', newline='') as infile:
            reader = csv.reader(infile)
            numRows = 0 
            for row in reader:
                if numRows == 0:
                    s = ''

                    for e in enumerate(row):
                        s = s + e[1]

                    headerHash = sha256(s.encode('utf-8')).hexdigest()
                    if (headerHash == "0599f2862ac011669d18fb17ecd7077bfa5e0b57584c3f7385fe0231d8b1e033"):
                        self._fileType = 'Anytone'
                        self._fileVersion = '3.4'
                    else:
                        # Starting wiht CPS 3.6 TxCC is added to column 55
                        if (headerHash == "4ef95deeedb33153d648c685a32d65777e64c234c32fc1459aeb821b1c9a82bb"):
                            self._fileType = 'Anytone'
                            self._fileVersion = '3.6'
                        else:
                            if (headerHash == "17a160a6bdc7c1bf759b4f61957d7753cc2373959b99d8b47087bbd9c44c0585"):
                                self._fileType = 'CS7000_analog_channels'
                                self._fileVersion = '9.00.93'
                            else:
                                if (headerHash == "68d829b43954be5483c42f92905d93fb6b8c7c47b28ef94bb51c7dd6a816f8d2"):
                                    self._fileType = 'CS7000_digital_channels'
                                    self._fileVersion = '9.00.93'
                                else:
                                    if (headerHash == "0f5e98e8c8aa9beb2dba3ad016070b300fb65b2c6cb2c5e6c4c8c2df7d1d9d4f"):
                                        self._fileType = 'CS7000'
                                    else:
                                        self._fileType = 'ERROR'

                numRows = numRows + 1

            self._rowsInFile = numRows

        infile.close()

    def LoadChannelNames(self, input_file):

        workbook = openpyxl.load_workbook(input_file)

        analogWorksheet = workbook["Analog Channels"]
        digitalWorksheet = workbook["Digital Channels"]

        inputs = []
        inputs.append(digitalWorksheet)

        if (self._includeAnalogChannels):
            inputs.append(analogWorksheet)

        processing_dmr_file = True
        for input_sheet in inputs:
            row = 2
            # cells(row, col) start counting with 1
            while (input_sheet.cell(row, 1).value != None):
                channelName = input_sheet.cell(row, 2).value
                channelSpacing = input_sheet.cell(row, 4).value
                rx_freq = input_sheet.cell(row, 12).value

                rx = Decimal(rx_freq)

                if processing_dmr_file == True:
                    tx_freq = input_sheet.cell(row, 18).value
                else:
                    tx_freq = input_sheet.cell(row, 19).value

                tx = Decimal(tx_freq)
                # Build list of UHF and VHF channel list
                if ( (rx >= 400) and (rx <= 512)) and ((tx >= 400) and (tx <=512)):
                    self._UhfChannels[channelName] = True
                else:
                    self._VhfChannels[channelName] = True
                row = row + 1

            processing_dmr_file = False

        workbook.close()

    def _SetArrays(self):

        # Header row for CS7000
        # Version 9.00.93 added "Contact Attribute" {Table, Direct} "Contact Type" {Private, Group, All}
        self._header_dmr = [
            "No.",
            "Channel Alias",
            "Digital ID [Per Each Channel]",
            "Color Code",
            "Slot Operation",
            "Scan/Roam/Vote List",
            "Auto Start Scan",
            "Rx Only",
            "Talk Around",
            "Lone Worker",
            "VOX",
            "Receive Frequency [MHz]",
            "Rx Ref Frequency",
            "Rx Group List",
            "Emergency Alarm Indication",
            "Emergency Alarm Ack",
            "Emergency Call Indication",
            "Transmit Frequency [MHz]",
            "Tx Ref Frequency",
            "Tx Contact Name",
            "Emergency System",
            "Power Level",
            "Tx Admit",
            "Tx Time-Out Time [s]",
            "TOT Re-Key Time [s]",
            "TOT Pre-Alert Time [s]",
            "Private Call Confirmed",
            "Data Call Confirmed",
            "Encryption",
            "Encryption Type",
            "Encryption Key List",
            "Pseudo Trunk",
            "Pseudo Trunk Designated TX",
            "Pilot_Freq Direct Mode",
            "Easy Trunking Mode",
            "Easy Trunking Sites",
            "TDMA Direct Mode",
            "Digital Channel Type",
            "IP Multi-site Connect",
            "Scan/Roam/Vote Mode",
            "Auto Start Roam",
            "Auto Start Voting",
            "In Call Criteria",
            "Text Message Format",
            "Encryption Ignore RX Clear Voice",
            "Compressed UDP Data Header",
            "Contact Attribute",
            "Contact Type"
        ]


        self._defaultRow_dmr = [
            "1",
            "DMR Smplx 441",
            "4",
            "1",
            "Slot 2",
            "None",
            "Off",
            "Off",
            "Off",
            "Off",
            "Off",
            "441",
            "Middle",
            "None",
            "Off",
            "Off",
            "Off",
            "441",
            "Middle",
            "None",
            "None",
            "High",
            "Always",
            "60",
            "0",
            "0",
            "Off",
            "Off",
            "Off",
            "Full",
            "Key 1",
            "Off",
            "Fixed",
            "Off",
            "Single-Repeater",
            "None",
            "Off",
            "Digital Channel",
            "Off",
            "None",
            "Off",
            "Off",
            "Follow Admit Criteria",
            "DMR Standard",
            "Off",
            "None",
            "Table",
            "Private"
        ]



        self._defaultRow_analog = [
            "1",
            "Channel 4",
            "Normal",
            "25.0",
            "Personality 1",
            "None",
            "Off",
            "Off",
            "Off",
            "Off",
            "Off",
            "441.00",
            "NONE",
            "NONE",
            "Middle",
            "CTCSS/CDCSS and Audio",
            "Carrier",
            "RX Squelch Mode",
            "441.00",
            "NONE",
            "NONE",
            "Middle",
            "High",
            "Aways Allow",
            "None",
            "Off",
            "60",
            "0",
            "0",
            "180",
            "None",
            "Off",
            "Off"
        ]


        self._header_analog = [
            "No.",
            "Channel Alias",
            "Carrier Squelch Level",
            "Channel Spacing [KHz]",
            "Personality List",
            "Scan/Roam/Vote List",
            "Auto Start Scan",
            "Rx Only",
            "Talk Around",
            "Lone Worker",
            "VOX",
            "Receive Frequency [MHz]",
            "Rx CTCSS/CDCSS Type",
            "CTCSS/CDCSS Type",
            "Rx Ref Frequency",
            "Rx Squelch Mode",
            "Monitor Squelch Mode",
            "Channel Switch Squelch Mode",
            "Transmit Frequency [MHz]",
            "Tx CTCSS/CDCSS Type",
            "CTCSS/CDCSS Type",
            "Tx Ref Frequency",
            "Power Level",
            "Tx Admit",
            "Emergency System",
            "CTCSS Tail Revert",
            "Tx Time-Out Time [s]",
            "TOT Re-Key Time [s]",
            "TOT Pre-Alert Time [s]",
            "CTCSS Tail Revert Option [Radians]",
            "Scan/Roam/Vote Mode",
            "Auto Start Roam",
            "Auto Start Voting"
        ]

        self._header_easyTrunkingChannels = [
            "No.",
            "Channel Alias",
            "Digital ID [Per Each Channel]",
            "Color Code",
            "Slot Operation",
            "Scan/Roam/Vote List",
            "Auto Start Scan",
            "Rx Only",
            "Talk Around",
            "Lone Worker",
            "VOX",
            "Receive Frequency [MHz]",
            "Rx Ref Frequency",
            "Rx Group List",
            "Emergency Alarm Indication",
            "Emergency Alarm Ack",
            "Emergency Call Indication",
            "Transmit Frequency [MHz]",
            "Tx Ref Frequency",
            "Tx Contact Name",
            "Emergency System",
            "Power Level",
            "Tx Admit",
            "Tx Time-Out Time [s]",
            "TOT Re-Key Time [s]",
            "TOT Pre-Alert Time [s]",
            "Private Call Confirmed",
            "Data Call Confirmed",
            "Encryption",
            "Encryption Type",
            "Encryption Key List",
            "Pseudo Trunk",
            "Pseudo Trunk Designated TX",
            "Pilot_Freq Direct Mode",
            "Easy Trunking Mode",
            "Easy Trunking Sites",
            "TDMA Direct Mode",
            "Digital Channel Type",
            "In Call Criteria",
            "Encryption Ignore RX Clear Voice"
        ]

        self._header_easyTrunkingPool = [
            "No",
            "Channel Alias",
            "Color Code",
            "Receive Frequency",
            "RX Ref Frequency",
            "Transmit Frequency",
            "TX Ref Frequency",
            "Slot Operation"
        ]

